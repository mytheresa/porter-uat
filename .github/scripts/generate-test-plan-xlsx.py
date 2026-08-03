import os
import sys
import json
import re
import math
from typing import Any, Optional, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REQUIRED_TOP_LEVEL_KEYS = [
    "EPIC_KEY",
    "EPIC_SLUG",
    "OUTPUT_PATH",
    "PLAN_TITLE",
    "GENERATED_DATE",
    "EPIC_SUMMARY",
    "EPIC_STATUS",
    "EPIC_CREATED_BY",
    "COMPONENT",
    "TARGET_URLS",
    "TIMEBOX",
    "COVERAGE_SUMMARY",
    "STORIES_IN_SCOPE",
    "STORIES_EXCLUDED",
    "OUT_OF_SCOPE",
    "DEV_EVIDENCE",
    "GAPS_SUMMARY",
    "CHECKLIST_ROWS",
    "MATRIX_ROWS",
    "EXPLORATORY_ROWS",
]

CHECKLIST_REQUIRED_FIELDS = ["Check ID", "Section", "Check", "How to Verify", "Pass Criteria"]
MATRIX_REQUIRED_FIELDS = [
    "Coverage ID",
    "Jira Source",
    "AC Ref",
    "Capability",
    "Priority",
    "Checklist Mapping",
    "AC Fidelity",
    "Evidence Notes",
    "Evidence Availability",
    "Case Type",
    "Inconsistencies",
]
EXPLORATORY_REQUIRED_FIELDS = [
    "Observation ID",
    "Source Type",
    "Jira Source",
    "Summary",
    "How to Validate",
    "Expected Observation",
    "Impact",
    "Linked Coverage IDs",
    "Evidence Notes",
]


def error_and_exit(message: str) -> None:
    print(f"Error: {message}")
    sys.exit(1)


def sanitize_for_excel(value: Any) -> Any:
    """
    Prevent formula injection by escaping values that Excel interprets as formulas.
    """
    if isinstance(value, str) and value and value[0] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def validate_top_level_schema(data: dict[str, Any]) -> None:
    missing = [k for k in REQUIRED_TOP_LEVEL_KEYS if k not in data]
    if missing:
        error_and_exit(f"Payload is missing required top-level keys: {', '.join(missing)}")


def validate_row_schema(section_name: str, rows: list[Any], required_fields: list[str]) -> None:
    if not isinstance(rows, list):
        error_and_exit(f"{section_name} must be a list.")

    for idx, row in enumerate(rows, start=1):
        if isinstance(row, dict):
            missing = [f for f in required_fields if f not in row]
            if missing:
                error_and_exit(
                    f"{section_name} row {idx} is missing required fields: {', '.join(missing)}"
                )
        elif isinstance(row, list):
            if len(row) < len(required_fields):
                error_and_exit(
                    f"{section_name} row {idx} has {len(row)} values; expected at least {len(required_fields)}."
                )
        else:
            error_and_exit(f"{section_name} row {idx} must be a dict or list.")


def validate_output_path(output_path: Any, epic_key: str) -> None:
    if not output_path:
        error_and_exit("OUTPUT_PATH must be provided and non-empty.")
    output_path_str = str(output_path)
    if os.path.isabs(output_path_str):
        error_and_exit("OUTPUT_PATH must be relative, not absolute.")

    normalized = os.path.normpath(output_path_str)
    expected_root = os.path.normpath("uat-test-plans")
    parent_dir = os.path.normpath(os.path.dirname(normalized))
    if parent_dir != expected_root:
        error_and_exit(
            f"OUTPUT_PATH must be directly inside '{expected_root}/'. Got '{normalized}'."
        )

    file_name = os.path.basename(normalized)
    if not file_name.startswith(f"{epic_key}-"):
        error_and_exit(
            f"OUTPUT_PATH filename must start with '{epic_key}-'. Got '{file_name}'."
        )
    if not normalized.endswith(".xlsx"):
        error_and_exit("OUTPUT_PATH must end with .xlsx")


def get_field(row: Any, headers: list[str], field_name: str) -> str:
    if isinstance(row, dict):
        return str(row.get(field_name, "")).strip()
    if isinstance(row, list):
        idx = headers.index(field_name)
        return str(row[idx]).strip() if idx < len(row) else ""
    return ""


def parse_mapping_ids(mapping_value: Any) -> list[str]:
    # Accept comma/semicolon separated mappings and trim whitespace.
    return [part.strip() for part in str(mapping_value).replace(";", ",").split(",") if part.strip()]


def validate_parity(checklist_rows: list[Any], matrix_rows: list[Any]) -> None:
    if not checklist_rows:
        error_and_exit("CHECKLIST_ROWS cannot be empty.")
    if not matrix_rows:
        error_and_exit("MATRIX_ROWS cannot be empty.")

    checklist_ids = []
    seen = set()
    for idx, row in enumerate(checklist_rows, start=1):
        cid = get_field(row, CHECKLIST_REQUIRED_FIELDS, "Check ID")
        if not cid:
            error_and_exit(f"CHECKLIST_ROWS row {idx} has empty 'Check ID'.")
        if cid in seen:
            error_and_exit(f"Duplicate Check ID found: '{cid}'.")
        seen.add(cid)
        checklist_ids.append(cid)

    checklist_id_set = set(checklist_ids)

    ac_refs = set()
    for idx, row in enumerate(matrix_rows, start=1):
        ac_ref = get_field(row, MATRIX_REQUIRED_FIELDS, "AC Ref")
        mapping = get_field(row, MATRIX_REQUIRED_FIELDS, "Checklist Mapping")
        if not ac_ref:
            error_and_exit(f"MATRIX_ROWS row {idx} has empty 'AC Ref'.")
        if not mapping:
            error_and_exit(f"MATRIX_ROWS row {idx} has empty 'Checklist Mapping'.")

        mapped_ids = parse_mapping_ids(mapping)
        if not mapped_ids:
            error_and_exit(f"MATRIX_ROWS row {idx} has invalid 'Checklist Mapping'.")

        for mapped_id in mapped_ids:
            if mapped_id not in checklist_id_set:
                error_and_exit(
                    f"MATRIX_ROWS row {idx} references unknown Check ID '{mapped_id}' in 'Checklist Mapping'."
                )
        ac_refs.add(ac_ref)

    if not ac_refs:
        error_and_exit("No AC references found in MATRIX_ROWS.")


def load_payload_or_exit(payload_path: str) -> dict[str, Any]:
    if not os.path.exists(payload_path):
        error_and_exit(f"Payload file not found at {payload_path}")

    try:
        with open(payload_path, "r", encoding="utf-8") as f:
            data: dict[str, Any] = json.load(f)
            return data
    except json.JSONDecodeError as e:
        error_and_exit(
            f"Malformed JSON in payload. Could not parse '{payload_path}'. Details: {e}"
        )
    except Exception as e:
        error_and_exit(f"Unexpected issue reading '{payload_path}'. Details: {e}")
    # Unreachable, but helps type checker
    return {}


def run_preflight_validations(data: dict[str, Any]) -> tuple[list[Any], list[Any], list[Any]]:
    validate_top_level_schema(data)

    epic_key: str = str(data.get("EPIC_KEY", "EPIC-KEY"))
    output_path: Any = data.get("OUTPUT_PATH")
    validate_output_path(output_path, epic_key)

    raw_checklist = data.get("CHECKLIST_ROWS", [])
    raw_matrix = data.get("MATRIX_ROWS", [])
    raw_exploratory = data.get("EXPLORATORY_ROWS", [])

    validate_row_schema("CHECKLIST_ROWS", raw_checklist, CHECKLIST_REQUIRED_FIELDS)
    validate_row_schema("MATRIX_ROWS", raw_matrix, MATRIX_REQUIRED_FIELDS)
    validate_row_schema("EXPLORATORY_ROWS", raw_exploratory, EXPLORATORY_REQUIRED_FIELDS)
    validate_parity(raw_checklist, raw_matrix)

    return raw_checklist, raw_matrix, raw_exploratory


def validate_payload(payload_path: str) -> None:
    data = load_payload_or_exit(payload_path)
    run_preflight_validations(data)
    print(f"Preflight validation passed: {payload_path}")


def build_token_telemetry(data: dict[str, Any], raw_checklist: list[Any], raw_matrix: list[Any], raw_exploratory: list[Any]) -> str:
    """
    Build a lightweight per-epic token telemetry string for the Overview tab.
    Estimation intentionally stays simple and cheap: tokens ~= chars / 4.
    """
    payload_json = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    payload_chars = len(payload_json)
    payload_bytes = len(payload_json.encode("utf-8"))
    est_tokens = math.ceil(payload_chars / 4)

    return (
        f"payload_bytes={payload_bytes}; payload_chars={payload_chars}; "
        f"est_tokens~{est_tokens}; checklist_rows={len(raw_checklist)}; "
        f"matrix_rows={len(raw_matrix)}; exploratory_rows={len(raw_exploratory)}"
    )

def standardize_row(row: Any, headers: list[str], enforce_values: Optional[dict[str, str]] = None) -> list[Any]:
    """
    Normalizes a row (dict or list) to match the exact length of headers.
    'enforce_values' is a dict mapping header names to mandatory values (e.g., {"Result": ""}).
    """
    if enforce_values is None:
        enforce_values = {}
        
    normalized = []
    if isinstance(row, dict):
        for header in headers:
            if header in enforce_values:
                normalized.append(enforce_values[header])
            else:
                normalized.append(row.get(header, ""))
                
    elif isinstance(row, list):
        for i, header in enumerate(headers):
            if header in enforce_values:
                normalized.append(enforce_values[header])
            elif i < len(row):
                normalized.append(str(row[i]))
            else:
                normalized.append("") 
    else:
        normalized = [""] * len(headers)
        
    return normalized


def standardize_checklist_row(row: Any, created_by: str) -> list[str]:
    """
    Normalize checklist rows into the new column order where Contact is placed
    right after Check ID. Contact is always derived from EPIC_CREATED_BY.
    """
    headers = ["Check ID", "Contact", "Section", "Check", "How to Verify", "Pass Criteria", "Result", "Notes"]

    def format_how_to_verify(value):
        """Ensure How to Verify is rendered as numbered multiline steps."""
        text = str(value or "").strip()
        if not text:
            return text

        # Keep already enumerated multiline content as-is.
        if "\n" in text and re.search(r"(^|\n)\s*\d+\.\s+", text):
            return text

        # If it's a single paragraph, split into sentence-like chunks and enumerate.
        if "\n" not in text:
            parts = [p.strip() for p in re.split(r"(?<=[.!?])\s+", text) if p.strip()]
            if len(parts) > 1:
                return "\n".join(f"{idx}. {part}" for idx, part in enumerate(parts, start=1))

        # Fallback: enumerate explicit lines or semicolon-separated steps.
        lines = [p.strip() for p in re.split(r"\n|;", text) if p.strip()]
        if len(lines) > 1:
            return "\n".join(f"{idx}. {line}" for idx, line in enumerate(lines, start=1))

        return text

    # Preferred shape: dict payload rows.
    if isinstance(row, dict):
        normalized = [
            row.get("Check ID", ""),
            created_by,
            row.get("Section", ""),
            row.get("Check", ""),
            format_how_to_verify(row.get("How to Verify", "")),
            row.get("Pass Criteria", ""),
            "",
            "",
        ]
        return normalized

    # Backward-compatible shape: old list rows without Contact.
    if isinstance(row, list):
        if len(row) >= 7:
            # Old workbook rows: [Check ID, Section, Check, How to Verify, Pass Criteria, Result, Notes]
            return [
                str(row[0]),
                created_by,
                str(row[1]),
                str(row[2]),
                format_how_to_verify(str(row[3])),
                str(row[4]),
                "",
                "",
            ]
        if len(row) >= 5:
            # Old minimal checklist rows: [Check ID, Section, Check, How to Verify, Pass Criteria]
            return [
                str(row[0]),
                created_by,
                str(row[1]),
                str(row[2]),
                format_how_to_verify(str(row[3])),
                str(row[4]),
                "",
                "",
            ]

    normalized = standardize_row(row, headers, {"Contact": created_by, "Result": "", "Notes": ""})
    if len(normalized) >= 5:
        normalized[4] = format_how_to_verify(normalized[4])
    return normalized

def create_uat_workbook(payload_path: str) -> None:
    data = load_payload_or_exit(payload_path)
    raw_checklist, raw_matrix, raw_exploratory = run_preflight_validations(data)
    telemetry_summary = build_token_telemetry(data, raw_checklist, raw_matrix, raw_exploratory)

    epic_key: str = str(data.get("EPIC_KEY", "EPIC-KEY"))
    output_path: str = str(data.get("OUTPUT_PATH", ""))

    wb = openpyxl.Workbook()
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws_overview = wb.active
    assert ws_overview is not None, "Failed to create Overview worksheet"
    ws_overview.title = "Overview"
    ws_overview.views.sheetView[0].showGridLines = True
    
    ws_overview["A1"] = data.get("PLAN_TITLE", f"{epic_key} UAT — Test Plan")
    ws_overview["A1"].font = title_font
    
    overview_data = [
        ("Epic Key", epic_key),
        ("Epic Slug", data.get("EPIC_SLUG", "")),
        ("Generated Date", data.get("GENERATED_DATE", "")),
        ("Epic Summary", data.get("EPIC_SUMMARY", "")),
        ("Epic Status", data.get("EPIC_STATUS", "")),
        ("Created By", data.get("EPIC_CREATED_BY", "")),
        ("Component", data.get("COMPONENT", "")),
        ("Target URLs", data.get("TARGET_URLS", "N/A")), 
        ("Estimated Timebox", data.get("TIMEBOX", "")),
        ("Coverage Summary", data.get("COVERAGE_SUMMARY", "")),
        ("Stories In Scope", data.get("STORIES_IN_SCOPE", "")),
        ("Stories Excluded", data.get("STORIES_EXCLUDED", "")),
        ("Out of Scope", data.get("OUT_OF_SCOPE", "")),
        ("Dev Evidence", data.get("DEV_EVIDENCE", "")),
        ("Gaps Summary", data.get("GAPS_SUMMARY", "")),
        ("Token Telemetry", telemetry_summary),
        ("Output Path", output_path)
    ]
    
    for row_idx, (label, val) in enumerate(overview_data, start=3):
        cell_lbl = ws_overview.cell(row=row_idx, column=1, value=label)
        
        if isinstance(val, list):
            val_str = ", ".join(str(v) for v in val)
        elif isinstance(val, dict):
            val_str = json.dumps(val)
        else:
            val_str = str(val) if val is not None else ""
            
        cell_val = ws_overview.cell(row=row_idx, column=2, value=sanitize_for_excel(val_str))
        
        cell_lbl.font = bold_font
        cell_val.font = regular_font
        cell_lbl.border = thin_border
        cell_val.border = thin_border

    def populate_sheet(ws, title_headers, rows_data):
        ws.views.sheetView[0].showGridLines = True
        
        # Style Headers
        for col_idx, header in enumerate(title_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 25
            
        # Populate Data & Dynamically Calculate Row Heights
        for row_idx, row_values in enumerate(rows_data, start=2):
            max_lines = 1
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=sanitize_for_excel(val))
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Estimate line breaks for explicit newlines + wrapped long text lines
                val_str = str(val or '')
                lines = val_str.split('\n')
                total_lines = 0
                for line in lines:
                    # Target width assumption max 45 printable chars per line block
                    total_lines += max(1, (len(line) // 45) + (1 if len(line) % 45 != 0 else 0))
                max_lines = max(max_lines, total_lines)
            
            # 15pt per line + 4pt padding to prevent clipping
            ws.row_dimensions[row_idx].height = max(18, (max_lines * 15) + 4)

    ws_checklist = wb.create_sheet(title="Checklist")
    checklist_headers = ["Check ID", "Contact", "Section", "Check", "How to Verify", "Pass Criteria", "Result", "Notes"]
    created_by = data.get("EPIC_CREATED_BY", "")
    processed_checklist = [standardize_checklist_row(r, created_by) for r in raw_checklist]
    populate_sheet(ws_checklist, checklist_headers, processed_checklist)

    ws_matrix = wb.create_sheet(title="Coverage Matrix")
    matrix_headers = ["Coverage ID", "Jira Source", "AC Ref", "Capability", "Priority", 
                      "Checklist Mapping", "AC Fidelity", "Evidence Notes", "Evidence Availability", "Case Type", "Inconsistencies"]
    processed_matrix = [standardize_row(r, matrix_headers) for r in raw_matrix]
    populate_sheet(ws_matrix, matrix_headers, processed_matrix)

    ws_exploratory = wb.create_sheet(title="Exploratory Scenarios")
    exploratory_headers = ["Observation ID", "Source Type", "Jira Source", "Summary", "How to Validate", 
                           "Expected Observation", "Impact", "Linked Coverage IDs", "Evidence Notes"]

    if not raw_exploratory:
        raw_exploratory = [["OBS-01", "N/A", "N/A", "None identified from explicit evidence", "N/A", "N/A", "Low", "N/A", "N/A"]]

    processed_exploratory = [standardize_row(r, exploratory_headers) for r in raw_exploratory]
    populate_sheet(ws_exploratory, exploratory_headers, processed_exploratory)

    # Auto-adjust Column Widths safely
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if ws.title == "Overview" and cell.coordinate == "A1":
                    continue
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                line_max = max(len(l) for l in lines) if lines else 0
                max_len = max(max_len, line_max)
            
            col_letter = get_column_letter(col[0].column if col[0].column else 1)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        
    wb.save(output_path)
    print(f"Successfully generated UAT test plan workbook at: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Error: Usage: python3 generate-test-plan-xlsx.py [--validate] <payload_path>")
        sys.exit(1)

    if sys.argv[1] == "--validate":
        if len(sys.argv) < 3:
            print("Error: Missing payload path for --validate mode.")
            sys.exit(1)
        payload_file = sys.argv[2]
        validate_payload(payload_file)
    else:
        payload_file = sys.argv[1]
        create_uat_workbook(payload_file)