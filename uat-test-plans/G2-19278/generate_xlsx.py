"""
UAT Test Plan XLSX Generator
============================
Canonical script for generating simplified UAT test plan workbooks.

Usage (agent-invoked via virtualenv):
    python3 -m venv /tmp/xlsx-venv
    /tmp/xlsx-venv/bin/pip install openpyxl -q
    /tmp/xlsx-venv/bin/python3 .github/scripts/generate-test-plan-xlsx.py

Rules enforced by this script (do not change without updating the skill):
- Header row: dark blue fill, white bold font — data rows: plain white
- No alternating row colours
- Result and Notes columns always empty (left for tester)
- No Status column in Coverage Matrix
- No separate brand coverage check row
- Overview includes Epic created by (from Jira creator field), NOT a "Created by" generator row
- Stories excluded (non-UI-testable) row always present after Stories in scope
- Column widths and row heights auto-fit from content

Agent instructions:
  1. Fill in all sections marked  ── AGENT: FILL ──
  2. Do not change the helper functions or styling constants
  3. Run with /tmp/xlsx-venv/bin/python3
  4. Output path follows: uat-test-plans/<EPIC_KEY>/<slug>.simplified-flows.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Styling constants (do not modify) ────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(size=10)
WRAP         = Alignment(wrap_text=True, vertical="top")
CENTER       = Alignment(horizontal="center", vertical="top", wrap_text=True)
THIN         = Side(style="thin", color="B8B8B8")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LINE_HEIGHT  = 15  # points per line at size 10


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = CENTER; c.border = BORDER


def style_row(ws, row, cols):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = BODY_FONT; c.alignment = WRAP; c.border = BORDER


def autofit(ws, min_col_width=8, max_col_width=60):
    """Auto-fit column widths and row heights from cell content."""
    col_widths = {}; row_lines = {}
    for row in ws.iter_rows():
        for cell in row:
            if not cell.value:
                continue
            lines = str(cell.value).split("\n")
            col = cell.column; r = cell.row
            longest = max(len(l) for l in lines)
            col_widths[col] = max(col_widths.get(col, min_col_width), longest)
            row_lines[r]    = max(row_lines.get(r, 1), len(lines))
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(w + 2, max_col_width)
    for r, lines in row_lines.items():
        ws.row_dimensions[r].height = max(LINE_HEIGHT, lines * LINE_HEIGHT)


# ════════════════════════════════════════════════════════════════════════════
# ── AGENT: FILL — Epic metadata ──────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════
EPIC_KEY        = "G2-19278"
EPIC_SLUG       = "plp-colour-linking"
PLAN_TITLE      = "G2-19278 UAT — PLP | Colour Linking"
GENERATED_DATE  = "2026-07-30"
EPIC_SUMMARY    = "G2-19278 — PLP | Colour linking"
EPIC_STATUS     = "Done"
EPIC_CREATED_BY = "Kozlova, Nika || External (nika.kozlova.external@mytheresa.com)"
COMPONENT       = "D1 — NAP/MRP - Search & Browse"
TIMEBOX         = "~60 minutes"
COVERAGE_SUMMARY = "8 AC-driven checks (CK-01–CK-08) + 3 Exploratory/Design Observations"
STORIES_IN_SCOPE = "G2-18440 — FE: PLP | Colour swatches (fully UI-testable)\nG2-18439 — BE: PLP API | Colour linking (mixed; BE pipeline items excluded, observable outcome covered by checks)"
STORIES_EXCLUDED = "G2-17553 — BE: Tapir | Indexing | Export colour linking data to Mink — pure backend data transfer, no observable UI outcome"
OUT_OF_SCOPE    = "PIM Data Ingestion (Discover3); Akamai/Madame swatch image flow; Sold-out swatch visual treatment on PLP (addressed separately)"
DEV_EVIDENCE    = "G2-18440: Available — PR #429 [G2-18440] PLP color swatches in atelier, MERGED 2026-03-12\nG2-18439: Available — PRs #2417 #2419 in hyena + PRs #45 #46 #47 #48 #49 in mink, all MERGED\nG2-19278 (Epic): Evidence Unavailable — no PRs/commits returned at Epic level"
GAPS_SUMMARY    = "GAP-01: Epic-level GitHub dev info returns no PRs (evidence attributed to story-level PRs)\nGAP-02: No explicit OOS visual spec on PLP; 'Sold out visibility — TBC' per G2-18440 Notes\nGAP-03: Test Scenarios in G2-18439 and G2-17553 both marked TBC; plan relies entirely on AC text\nGAP-04: Bug G2-20786 referenced in G2-18440 comment; resolution status unverified — confirm before UAT"

OUTPUT_PATH     = f"uat-test-plans/{EPIC_KEY}/{EPIC_SLUG}.simplified-flows.xlsx"


# ── AGENT: FILL — Sheet 1: Checklist rows ─────────────────────────────────────
# Format: (check_id, section, check, how_to_verify, pass_criteria, "", "")
# Result (col 6) and Notes (col 7) are ALWAYS empty strings — never pre-populate
CHECKLIST_ROWS = [
    ("CK-01", "1 — Swatch Display",
     "Colour swatches are displayed below the product tile on PLP",
     "1. Navigate to acceptance.net-a-porter.com/en-de or acceptance.mrporter.com/en-de.\n2. Open a category or search results page.\n3. Locate a product with multiple colour variants.\n4. Observe the area below the product image tile.",
     "Swatches visible below product tile for colour-linked products.", "", "Maps to G2-19278_TC01 | G2-18440 AC1"),
    ("CK-02", "2 — Swatch Interaction",
     "Clicking a swatch swaps the main product image, updates URL and highlights the selected swatch",
     "1. Find a product with multiple swatches on PLP.\n2. Note the initial product image and page URL.\n3. Click a swatch different from the default colour.\n4. Observe image, URL and swatch visual state.",
     "Image swaps to selected colour; URL updates to new slug; clicked swatch visually highlighted.", "", "Maps to G2-19278_TC02 | G2-18440 AC2"),
    ("CK-03", "3 — Overflow Colours",
     "Maximum 4 swatches shown; (+) indicator present for products with 5+ colour variants",
     "1. Find a product with 5 or more colour variants.\n2. Count swatches displayed below the tile.\n3. Check for a (+) overflow indicator.",
     "Exactly 4 swatches displayed; (+) visible for products with more than 4 colours.", "", "Maps to G2-19278_TC03 | G2-18440 AC3"),
    ("CK-04", "3 — Overflow Colours",
     "Hovering over (+) restores the cover/main product image",
     "1. Click a non-default swatch to change the tile image.\n2. Move cursor to hover over the (+) indicator.\n3. Observe the product tile image.",
     "Cover/main product image restored on (+) hover; no navigation triggered.", "", "Maps to G2-19278_TC04 | G2-18440 AC4"),
    ("CK-05", "3 — Overflow Colours",
     "Clicking (+) navigates to the cover/default colour PDP",
     "1. On a tile showing (+), click the (+) indicator.\n2. Observe the resulting page and URL.",
     "Browser navigates to the PDP of the cover/default colour variant.", "", "Maps to G2-19278_TC05 | G2-18440 AC5"),
    ("CK-06", "4 — Stock Visibility",
     "Out of stock colour variants remain visible as swatches on PLP",
     "1. Identify a product with at least one OOS colour variant (coordinate with trading/QA for test data).\n2. Navigate to a PLP listing that product.\n3. Observe the swatch row below the tile.",
     "OOS colour swatch visible below tile; not suppressed.", "", "Maps to G2-19278_TC06 | G2-18440 AC6 | Note: GAP-02 — no explicit OOS visual spec"),
    ("CK-07", "5 — Deep Linking",
     "Each swatch click navigates to that colour's correct PDP URL",
     "1. On a multi-colour product tile, click each swatch in turn.\n2. Check or copy the resulting URL after each click.\n3. Verify each URL matches the expected slug for that colour variant.",
     "Every swatch resolves to its correct colour-specific PDP URL.", "", "Maps to G2-19278_TC07 | G2-18440 AC7"),
    ("CK-08", "6 — Dual Brand Coverage",
     "Colour swatches feature functional on both NAP and MRP",
     "1. Repeat CK-01 to CK-07 on acceptance.net-a-porter.com/en-de.\n2. Repeat CK-01 to CK-07 on acceptance.mrporter.com/en-de.",
     "All swatch checks pass on both NAP and MRP.", "", "Maps to G2-19278_TC08 | Epic AC Scope"),
]


# ── AGENT: FILL — Sheet 3: Coverage Matrix rows ───────────────────────────────
# Format: (coverage_id, jira_source, ac_ref, capability, priority,
#          checklist_mapping, ac_fidelity, evidence_notes, evidence_availability, case_type)
# No Status column.
MATRIX_ROWS = [
    ("G2-19278_TC01", "G2-18440", "AC1",
     "Colour swatches displayed below product tile on PLP",
     "High", "CK-01", "Exact",
     "PR #429 [G2-18440] PLP color swatches — atelier — MERGED 2026-03-12",
     "Available", "AC"),
    ("G2-19278_TC02", "G2-18440", "AC2",
     "Selecting a swatch swaps main product image, updates PLP URL and visually highlights selected swatch",
     "High", "CK-02", "Exact",
     "PR #429 [G2-18440] PLP color swatches — atelier — MERGED 2026-03-12",
     "Available", "AC"),
    ("G2-19278_TC03", "G2-18440", "AC3",
     "Maximum 4 swatches shown on tile; (+) indicator displayed when more colours exist",
     "High", "CK-03", "Exact",
     "PR #429 [G2-18440] PLP color swatches — atelier — MERGED 2026-03-12",
     "Available", "AC"),
    ("G2-19278_TC04", "G2-18440", "AC4",
     "Hovering over (+) restores the cover / main product image",
     "Medium", "CK-04", "Exact",
     "PR #429 [G2-18440] PLP color swatches — atelier — MERGED 2026-03-12",
     "Available", "AC"),
    ("G2-19278_TC05", "G2-18440", "AC5",
     "Clicking (+) navigates user to the cover / default colour PDP",
     "Medium", "CK-05", "Exact",
     "PR #429 [G2-18440] PLP color swatches — atelier — MERGED 2026-03-12",
     "Available", "AC"),
    ("G2-19278_TC06", "G2-18440", "AC6",
     "Out of stock colour variants remain visible as swatches on PLP (no suppression)",
     "Medium", "CK-06", "Exact",
     "PR #429 — atelier MERGED; bug G2-20786 noted in comments (verify resolution before UAT)",
     "Available", "AC"),
    ("G2-19278_TC07", "G2-18440", "AC7",
     "Deep link: each swatch click navigates to that colour's correct PDP URL",
     "High", "CK-07", "Exact",
     "PR #429 [G2-18440] PLP color swatches — atelier — MERGED 2026-03-12",
     "Available", "AC"),
    ("G2-19278_TC08", "G2-19278 (Epic) / G2-18440", "Epic AC1 + Scope",
     "Colour swatches feature fully functional on both NAP and MRP brands",
     "High", "CK-08", "Exact",
     "Epic AC states both NAP and MRP; G2-18440 explicitly scoped to both brands; Epic-level GitHub: Evidence Unavailable",
     "Evidence Unavailable (Epic-level GitHub)", "AC"),
]


# ── AGENT: FILL — Sheet 4: Exploratory observations ──────────────────────────
# Format: (obs_id, source_type, jira_source, summary, how_to_validate,
#          expected_observation, impact, linked_coverage_ids, evidence_notes, status)
# status is always "OBSERVE"
# If no observations: one row with ("OBS-01", "N/A", "N/A", "None identified from explicit evidence", ...)
EXPLORATORY_ROWS = [
    ("OBS-01", "Comment", "G2-18440 (comment 738669)",
     "Bug G2-20786 raised for some products' colour swatches; resolution status unknown at plan generation time",
     "Identify products affected by G2-20786; check swatch display and interaction for correctness",
     "Swatch display and interaction consistent across all products; no broken or missing swatches",
     "Medium — could affect conversion on specific SKUs",
     "G2-19278_TC01, G2-19278_TC02",
     "G2-20786 referenced by ext.pjassu in G2-18440 comment 738669; verify status before UAT sign-off",
     "OBSERVE"),
    ("OBS-02", "Story Note", "G2-18439 (comment 734636)",
     "A product should NOT appear in its own alternativeColors list; backend changed to exclude self-references",
     "On PLP, verify swatches do not include the cover colour as a duplicate; cover is either the active default or absent from the overflow row",
     "No duplicate self-reference swatch; swatch count accurate to actual colour variant count",
     "Low — data quality; affects user perception of colour count",
     "G2-19278_TC03",
     "Confirmed by ext.pjassu comment on G2-18439; PRs #2417 #2419 in hyena MERGED with this fix",
     "OBSERVE"),
    ("OBS-03", "Design", "G2-18440 (description)",
     "Figma designs provided for NAP and MRP; swatch layout and style should match brand designs",
     "Compare rendered swatches on acceptance against Figma — NAP: https://www.figma.com/design/AClD0sJnAeZlWbchDF8JHm | MRP: https://www.figma.com/design/UDgQqNHggMzHOTqaBaZ37a — check swatch size, spacing, highlight style and (+) indicator per brand",
     "Swatch size, spacing, highlight style and (+) overflow indicator align with Figma designs for each brand",
     "Medium — visual inconsistency may reduce brand confidence",
     "G2-19278_TC01, G2-19278_TC08",
     "Figma links explicitly in G2-18440 description; screenshot of expected layout embedded in story",
     "OBSERVE"),
]


# ════════════════════════════════════════════════════════════════════════════
# ── Workbook generation (do not modify below this line) ──────────────────────
# ════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# Sheet 1: Checklist
ws1 = wb.worksheets[0]
ws1.title = "Checklist"
ws1.freeze_panes = "A2"
checklist_headers = ["Check ID", "Section", "Check", "How to Verify", "Pass Criteria", "Result", "Notes"]
for col, h in enumerate(checklist_headers, 1):
    ws1.cell(row=1, column=col, value=h)
style_header(ws1, 1, len(checklist_headers))
for r_idx, row_data in enumerate(CHECKLIST_ROWS, 2):
    for c_idx, val in enumerate(row_data, 1):
        ws1.cell(row=r_idx, column=c_idx, value=val)
    style_row(ws1, r_idx, len(checklist_headers))
autofit(ws1)

# Sheet 2: Overview
ws2 = wb.create_sheet("Overview")
overview_rows = [
    ("Plan",                               PLAN_TITLE),
    ("Generated",                          GENERATED_DATE),
    ("Epic",                               EPIC_SUMMARY),
    ("Epic Status",                        EPIC_STATUS),
    ("Epic created by",                    EPIC_CREATED_BY),
    ("Component",                          COMPONENT),
    ("Application URL — NAP",             "https://acceptance.net-a-porter.com/en-de"),
    ("Application URL — MRP",             "https://acceptance.mrporter.com/en-de"),
    ("Audience",                           "Business quick-check"),
    ("Timebox",                            TIMEBOX),
    ("Coverage",                           COVERAGE_SUMMARY),
    ("Stories in scope",                   STORIES_IN_SCOPE),
    ("Stories excluded (non-UI-testable)", STORIES_EXCLUDED),
    ("Out of scope",                       OUT_OF_SCOPE),
    ("Dev Evidence",                       DEV_EVIDENCE),
    ("Gaps",                               GAPS_SUMMARY),
]
for col, h in enumerate(["Field", "Value"], 1):
    ws2.cell(row=1, column=col, value=h)
style_header(ws2, 1, 2)
for r_idx, (field, val) in enumerate(overview_rows, 2):
    ws2.cell(row=r_idx, column=1, value=field)
    ws2.cell(row=r_idx, column=2, value=val)
    style_row(ws2, r_idx, 2)
autofit(ws2)

# Sheet 3: Coverage Matrix
ws3 = wb.create_sheet("Coverage Matrix")
ws3.freeze_panes = "A2"
matrix_headers = [
    "Coverage ID", "Jira Source", "AC Ref", "Capability / Requirement",
    "Priority", "Checklist Mapping",
    "AC Fidelity", "Evidence Notes", "Evidence Availability", "Case Type",
]
for col, h in enumerate(matrix_headers, 1):
    ws3.cell(row=1, column=col, value=h)
style_header(ws3, 1, len(matrix_headers))
for r_idx, row_data in enumerate(MATRIX_ROWS, 2):
    for c_idx, val in enumerate(row_data, 1):
        ws3.cell(row=r_idx, column=c_idx, value=val)
    style_row(ws3, r_idx, len(matrix_headers))
autofit(ws3)

# Sheet 4: Exploratory and Design Obs
ws4 = wb.create_sheet("Exploratory and Design Obs")
ws4.freeze_panes = "A2"
obs_headers = [
    "Observation ID", "Source Type", "Jira Source", "Observation Summary",
    "How to Validate Manually", "Expected Observation", "Impact",
    "Linked Coverage IDs", "Evidence Notes", "Status",
]
for col, h in enumerate(obs_headers, 1):
    ws4.cell(row=1, column=col, value=h)
style_header(ws4, 1, len(obs_headers))
for r_idx, row_data in enumerate(EXPLORATORY_ROWS, 2):
    for c_idx, val in enumerate(row_data, 1):
        ws4.cell(row=r_idx, column=c_idx, value=val)
    style_row(ws4, r_idx, len(obs_headers))
autofit(ws4)

# Save
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
wb.save(OUTPUT_PATH)
print(f"Saved: {OUTPUT_PATH}")
